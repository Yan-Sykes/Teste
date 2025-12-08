import time
import psutil
import subprocess
import win32com.client
import os
import openpyxl
import pythoncom

# ========================================
# ⚙️ CONFIGURAÇÕES GERAIS DO SISTEMA
# ========================================
# Nome da conexão SAP utilizada para acesso ao sistema ERP
SAP_NOME_CONEXAO = "s10   P11   ERP-Prod"

# Diretório de rede onde os arquivos exportados serão salvos
CAMINHO_EXPORTACAO = r"\\br03file\pcoudir\Operacoes\10. Planning Raw Material\Gerenciamento de materiais\Monitor de validades"

# Nomes dos arquivos de exportação do SAP
ARQUIVO_MB51 = "Mb51_SAP.xlsx"      # Relatório de movimentações de material
ARQUIVO_SQ00 = "Sq00_Validade.xlsx"  # Relatório de validades
ARQUIVO_VENC = "Vencimentos_SAP.xlsx"  # Relatório de vencimentos

# ========================================
# 🧹 GERENCIAMENTO DE PROCESSOS SAP
# ========================================
def verificar_e_fechar_sap():
    """
    Verifica se o SAP está em execução e encerra todos os processos relacionados.
    
    Esta função é necessária para garantir que não haja conflitos ao abrir
    uma nova sessão do SAP para exportação de dados.
    
    Returns:
        bool: True se o SAP estava aberto e foi fechado, False caso contrário
        
    Raises:
        Exception: Captura e registra erros durante a verificação/fechamento
    """
    try:
        sap_aberto = False
        # Itera por todos os processos do sistema procurando por processos SAP
        for proc in psutil.process_iter(['name']):
            if proc.info['name'] and 'sap' in proc.info['name'].lower():
                sap_aberto = True
                proc.terminate()  # Encerra o processo SAP
        
        if sap_aberto:
            print("🧹 SAP estava aberto - fechando...")
            time.sleep(5)  # Aguarda o encerramento completo dos processos
        else:
            print("SAP não estava aberto")
        return sap_aberto
    except Exception as e:
        print(f"❌ Erro ao verificar/fechar SAP: {e}")
        return False


# ========================================
# 📂 GERENCIAMENTO DE ARQUIVOS EXCEL
# ========================================
def fechar_arquivo_excel(nome_arquivo):
    """
    Fecha um arquivo Excel específico que esteja aberto no sistema.
    
    Esta função utiliza a API COM do Windows para conectar-se a instâncias
    ativas do Excel e fechar arquivos específicos. É especialmente útil para
    fechar arquivos que foram abertos automaticamente pelo SAP durante a exportação.
    
    Estratégia de fechamento:
    1. Tenta conectar usando Dispatch (instância ativa)
    2. Se falhar, tenta GetObject (instância existente)
    3. Itera pelos workbooks abertos procurando o arquivo
    4. Fecha o arquivo sem salvar alterações
    5. Fecha o Excel se não houver mais arquivos abertos
    
    Args:
        nome_arquivo (str): Nome do arquivo Excel (ex: "Mb51_SAP.xlsx")
    
    Returns:
        bool: True se o arquivo foi fechado com sucesso, False caso contrário
        
    Note:
        Requer pythoncom para inicialização COM em ambientes multi-thread
    """
    try:
        # Inicializa o ambiente COM para comunicação com aplicações Windows
        pythoncom.CoInitialize()
        
        excel_app = None
        arquivo_fechado = False
        
        # Estratégia 1: Tenta conectar usando Dispatch (instância ativa do Excel)
        try:
            excel_app = win32com.client.Dispatch("Excel.Application")
            print(f"📂 Excel encontrado com {excel_app.Workbooks.Count} arquivo(s) aberto(s)")
        except:
            # Estratégia 2: Se Dispatch falhar, tenta GetObject (instância existente)
            try:
                excel_app = win32com.client.GetObject(None, "Excel.Application")
                print(f"📂 Excel encontrado (GetObject) com {excel_app.Workbooks.Count} arquivo(s) aberto(s)")
            except:
                # Excel não está em execução - arquivo já deve estar fechado
                print(f"ℹ️ Excel não está aberto. Arquivo '{nome_arquivo}' já deve estar fechado.")
                return True
        
        if excel_app:
            # Itera por todos os workbooks abertos procurando o arquivo específico
            for wb in excel_app.Workbooks:
                print(f"   🔍 Verificando: {wb.Name}")
                if wb.Name == os.path.basename(nome_arquivo):
                    # Fecha o arquivo sem salvar alterações (SaveChanges=False)
                    wb.Close(SaveChanges=False)
                    print(f"✅ Arquivo '{nome_arquivo}' fechado com sucesso.")
                    arquivo_fechado = True
                    break
            
            if not arquivo_fechado:
                print(f"⚠️ Arquivo '{nome_arquivo}' não foi encontrado entre os arquivos abertos.")
            
            # Otimização: Fecha o Excel completamente se não houver mais arquivos abertos
            if excel_app.Workbooks.Count == 0:
                excel_app.Quit()
                print("🔒 Excel fechado completamente (não havia outros arquivos abertos).")
        
        return arquivo_fechado
        
    except Exception as e:
        print(f"❌ Erro ao fechar '{nome_arquivo}': {e}")
        return False
    finally:
        # Sempre finaliza o ambiente COM para liberar recursos
        pythoncom.CoUninitialize()


def fechar_mb51():
    """
    Fecha o arquivo de relatório MB51 (movimentações de material).
    
    Returns:
        bool: True se o arquivo foi fechado com sucesso
    """
    return fechar_arquivo_excel(ARQUIVO_MB51)


def fechar_sq00():
    """
    Fecha o arquivo de relatório SQ00 (validades).
    
    Returns:
        bool: True se o arquivo foi fechado com sucesso
    """
    return fechar_arquivo_excel(ARQUIVO_SQ00)


def fechar_venc():
    """
    Fecha o arquivo de relatório de vencimentos.
    
    Returns:
        bool: True se o arquivo foi fechado com sucesso
    """
    return fechar_arquivo_excel(ARQUIVO_VENC)


def forcar_fechar_excel():
    """
    Força o encerramento de todos os processos do Excel no sistema.
    
    Esta função é utilizada como último recurso quando o método COM falha
    em fechar arquivos específicos. Encerra todos os processos Excel.exe
    encontrados no sistema operacional.
    
    Returns:
        bool: True se processos foram encerrados, False caso contrário
        
    Warning:
        Esta função fecha TODOS os processos Excel, incluindo arquivos
        que o usuário possa ter aberto manualmente. Use com cautela.
    """
    try:
        excel_fechado = False
        # Itera por todos os processos procurando por Excel
        for proc in psutil.process_iter(['name']):
            if proc.info['name'] and 'excel' in proc.info['name'].lower():
                proc.terminate()  # Encerra o processo Excel
                excel_fechado = True
        
        if excel_fechado:
            print("🔨 Excel foi forçado a fechar.")
            time.sleep(3)  # Aguarda o encerramento completo dos processos
        else:
            print("ℹ️ Excel não estava em execução.")
        
        return True
    except Exception as e:
        print(f"❌ Erro ao forçar fechamento do Excel: {e}")
        return False


def aguardar_arquivo_disponivel(caminho_arquivo, timeout=60):
    """
    Aguarda até que um arquivo esteja disponível para leitura/escrita.
    
    Esta função é útil após exportações do SAP, pois os arquivos podem
    permanecer bloqueados por alguns segundos enquanto o Excel finaliza
    a gravação. A função tenta abrir o arquivo periodicamente até que
    esteja disponível ou o timeout seja atingido.
    
    Args:
        caminho_arquivo (str): Caminho completo do arquivo a verificar
        timeout (int): Tempo máximo de espera em segundos (padrão: 60)
    
    Returns:
        bool: True se o arquivo ficou disponível, False se timeout foi atingido
        
    Note:
        A função verifica a disponibilidade tentando abrir o arquivo em
        modo append, o que requer permissões de escrita.
    """
    print(f"⏳ Aguardando arquivo '{os.path.basename(caminho_arquivo)}' ficar disponível...")
    tempo_inicial = time.time()
    
    while True:
        # Verifica se o arquivo existe no sistema de arquivos
        if os.path.exists(caminho_arquivo):
            try:
                # Tenta abrir o arquivo em modo append para verificar disponibilidade
                # Se conseguir abrir, o arquivo não está bloqueado
                with open(caminho_arquivo, "a"):
                    print(f"✅ Arquivo '{os.path.basename(caminho_arquivo)}' disponível!")
                    return True
            except (PermissionError, OSError):
                # Arquivo ainda está bloqueado por outro processo
                pass
        
        # Verifica se o tempo limite foi atingido
        if time.time() - tempo_inicial > timeout:
            print(f"⚠️ Timeout aguardando '{os.path.basename(caminho_arquivo)}'")
            return False
        
        # Aguarda 1 segundo antes de tentar novamente
        time.sleep(1)


# ========================================
# 🔐 CONEXÃO E AUTENTICAÇÃO SAP
# ========================================
def abrir_sap_e_fazer_logon():
    """
    Abre o SAP Logon e estabelece uma sessão de scripting.
    
    Esta função realiza as seguintes etapas:
    1. Inicia o aplicativo SAP Logon
    2. Aguarda a inicialização completa
    3. Conecta ao Scripting Engine do SAP
    4. Abre uma conexão com o sistema configurado
    5. Obtém a sessão ativa para automação
    
    Returns:
        session: Objeto de sessão SAP para automação, ou None em caso de erro
        
    Raises:
        Exception: Captura e registra erros durante o processo de conexão
        
    Note:
        Requer que o SAP GUI Scripting esteja habilitado no sistema.
        O nome da conexão deve estar configurado em SAP_NOME_CONEXAO.
    """
    try:
        print("🔐 Abrindo SAP Logon...")
        # Inicia o aplicativo SAP Logon
        subprocess.Popen([r"C:\Program Files (x86)\SAP\FrontEnd\SAPgui\saplogon.exe"])
        time.sleep(7)  # Aguarda inicialização completa do SAP

        # Obtém referência ao objeto SAPGUI para automação
        sap_gui_auto = win32com.client.GetObject("SAPGUI")
        if not sap_gui_auto:
            raise Exception("SAPGUI não disponível")

        # Obtém o Scripting Engine para automação de transações
        application = sap_gui_auto.GetScriptingEngine
        if not application:
            raise Exception("Scripting Engine não disponível")

        # Abre conexão com o sistema SAP configurado
        print(f"🔗 Conectando à entrada do SAP Logon: '{SAP_NOME_CONEXAO}'...")
        connection = application.OpenConnection(SAP_NOME_CONEXAO, True)
        if not connection:
            raise Exception(f"Conexão '{SAP_NOME_CONEXAO}' não foi encontrada.")

        # Obtém a sessão ativa para executar comandos
        session = connection.Children(0)
        if not session:
            raise Exception("Sessão SAP não encontrada.")

        # Envia tecla Enter para confirmar login (se necessário)
        session.findById("wnd[0]").sendVKey(0)
        print("✅ SAP aberto e logon iniciado com sucesso!")
        return session
    except Exception as e:
        print(f"❌ Erro ao abrir SAP: {e}")
        return None


# ========================================
# 📦 TRANSAÇÕES SAP - EXPORTAÇÃO DE DADOS
# ========================================
def executar_mb51(session):
    """
    Executa a transação MB51 (Documento de Material) e exporta os dados para Excel.
    
    A transação MB51 fornece informações sobre movimentações de materiais,
    incluindo entradas, saídas, transferências e outros movimentos de estoque.
    
    Sequência de automação:
    1. Maximiza a janela SAP
    2. Navega para a transação MB51
    3. Carrega layout salvo com filtros pré-configurados
    4. Executa a consulta
    5. Exporta os resultados para Excel no caminho especificado
    
    Args:
        session: Objeto de sessão SAP ativa
        
    Raises:
        Exception: Captura e registra erros durante a execução da transação
        
    Note:
        A transação deve ter um layout salvo com os filtros necessários.
        O arquivo será exportado para CAMINHO_EXPORTACAO/ARQUIVO_MB51.
    """
    try:
        print("📊 Executando sequência MB51...")
        s = session

        # Maximiza a janela principal do SAP
        session.findById("wnd[0]").maximize()
        
        # Navega para a transação MB51
        session.findById("wnd[0]/tbar[0]/okcd").text = "/nMB51"
        session.findById("wnd[0]").sendVKey(0)
        time.sleep(2)  # Aguarda carregamento da transação

        # Carrega layout salvo e executa a consulta
        s.findById("wnd[0]/tbar[1]/btn[17]").press()  # Botão de layout
        s.findById("wnd[1]/tbar[0]/btn[8]").press()   # Confirma seleção
        s.findById("wnd[0]/tbar[1]/btn[8]").press()   # Executa consulta

        # Configura colunas para exportação
        s.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").setCurrentCell(14, "EBELN")
        s.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").selectedRows = "14"
        s.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").contextMenu()

        s.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").setCurrentCell(20, "LGORT")
        s.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").selectedRows = "20"
        s.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").contextMenu()
        
        # Inicia exportação para Excel
        s.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").selectContextMenuItem("&XXL")

        # Configura caminho e nome do arquivo de exportação
        s.findById("wnd[1]/tbar[0]/btn[0]").press()
        s.findById("wnd[1]/usr/ctxtDY_PATH").text = CAMINHO_EXPORTACAO
        s.findById("wnd[1]/usr/ctxtDY_FILENAME").text = ARQUIVO_MB51
        s.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = len(ARQUIVO_MB51)
        s.findById("wnd[1]/tbar[0]/btn[11]").press()  # Confirma exportação

        time.sleep(2)  # Aguarda conclusão da exportação
        print(f"✅ MB51 exportado para {os.path.join(CAMINHO_EXPORTACAO, ARQUIVO_MB51)}")

    except Exception as e:
        print(f"❌ Erro na execução MB51: {e}")


def executar_sq00(session):
    """
    Executa a transação SQ00 (Query SAP) e exporta dados de validade para Excel.
    
    A transação SQ00 permite executar queries personalizadas no SAP.
    Esta função executa uma query específica que retorna informações sobre
    validades de materiais em estoque.
    
    Sequência de automação:
    1. Maximiza a janela SAP
    2. Navega para a transação SQ00
    3. Seleciona a query salva
    4. Executa a query com parâmetros pré-configurados
    5. Exporta os resultados para Excel
    
    Args:
        session: Objeto de sessão SAP ativa
        
    Raises:
        Exception: Captura e registra erros durante a execução da query
        
    Note:
        A query deve estar previamente configurada no sistema SAP.
        O arquivo será exportado para CAMINHO_EXPORTACAO/ARQUIVO_SQ00.
    """
    try:
        print("📊 Executando sequência SQ00...")
        s = session

        # Maximiza a janela principal do SAP
        session.findById("wnd[0]").maximize()
        
        # Navega para a transação SQ00 (Query SAP)
        session.findById("wnd[0]/tbar[0]/okcd").text = "/nSQ00"
        session.findById("wnd[0]").sendVKey(0)
        time.sleep(2)  # Aguarda carregamento da transação

        # Abre lista de queries disponíveis
        s.findById("wnd[0]/tbar[1]/btn[19]").press()

        # Seleciona a query específica (linha 4)
        s.findById("wnd[1]/usr/cntlGRID1/shellcont/shell").currentCellRow = 4
        s.findById("wnd[1]/usr/cntlGRID1/shellcont/shell").selectedRows = "4"
        s.findById("wnd[1]/usr/cntlGRID1/shellcont/shell").doubleClickCurrentCell()

        # Navega pelos parâmetros da query
        s.findById("wnd[0]/usr/cntlGRID_CONT0050/shellcont/shell").currentCellRow = 52
        s.findById("wnd[0]/usr/cntlGRID_CONT0050/shellcont/shell").firstVisibleRow = 20
        s.findById("wnd[0]/usr/cntlGRID_CONT0050/shellcont/shell").selectedRows = "52"
        
        # Executa a query
        s.findById("wnd[0]/tbar[1]/btn[8]").press()
        
        # Carrega layout salvo
        s.findById("wnd[0]/tbar[1]/btn[17]").press()
        s.findById("wnd[1]/tbar[0]/btn[8]").press()

        # Confirma seleção de layout
        s.findById("wnd[1]/usr/cntlALV_CONTAINER_1/shellcont/shell").selectedRows = "0"
        s.findById("wnd[1]/usr/cntlALV_CONTAINER_1/shellcont/shell").doubleClickCurrentCell()
        s.findById("wnd[0]/tbar[1]/btn[8]").press()

        # Prepara exportação para Excel
        s.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").setCurrentCell(9, "TEXT_MCHB_MATNR")
        s.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").selectedRows = "9"
        s.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").contextMenu()
        s.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").selectContextMenuItem("&XXL")

        # Configura caminho e nome do arquivo de exportação
        s.findById("wnd[1]/tbar[0]/btn[0]").press()
        s.findById("wnd[1]/usr/ctxtDY_PATH").text = CAMINHO_EXPORTACAO
        s.findById("wnd[1]/usr/ctxtDY_FILENAME").text = ARQUIVO_SQ00
        s.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = len(ARQUIVO_SQ00)
        s.findById("wnd[1]/tbar[0]/btn[11]").press()  # Confirma exportação

        time.sleep(2)  # Aguarda conclusão da exportação
        print(f"✅ SQ00 exportado para {os.path.join(CAMINHO_EXPORTACAO, ARQUIVO_SQ00)}")

    except Exception as e:
        print(f"❌ Erro na execução SQ00: {e}")


def executar_sq00_venc(session):
    try:
        print("📊 Executando sequência SQ00 para Vencimentos...")
        s = session

        session.findById("wnd[0]").maximize()
        session.findById("wnd[0]/tbar[0]/okcd").text = "/nSQ00"
        session.findById("wnd[0]").sendVKey(0)
        time.sleep(2)

        s.findById("wnd[0]/tbar[1]/btn[19]").press()

        s.findById("wnd[1]/usr/cntlGRID1/shellcont/shell").currentCellRow = 4
        s.findById("wnd[1]/usr/cntlGRID1/shellcont/shell").selectedRows = "4"
        s.findById("wnd[1]/usr/cntlGRID1/shellcont/shell").doubleClickCurrentCell()

        s.findById("wnd[0]/usr/cntlGRID_CONT0050/shellcont/shell").currentCellRow = 52
        s.findById("wnd[0]/usr/cntlGRID_CONT0050/shellcont/shell").firstVisibleRow = 20
        s.findById("wnd[0]/usr/cntlGRID_CONT0050/shellcont/shell").selectedRows = "52"
        s.findById("wnd[0]/tbar[1]/btn[8]").press()
        s.findById("wnd[0]/tbar[1]/btn[17]").press()
        s.findById("wnd[1]/tbar[0]/btn[8]").press()

        s.findById("wnd[1]/usr/cntlALV_CONTAINER_1/shellcont/shell").selectedRows = "0"
        s.findById("wnd[1]/usr/cntlALV_CONTAINER_1/shellcont/shell").doubleClickCurrentCell()
        s.findById("wnd[0]/tbar[1]/btn[8]").press()

        s.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").setCurrentCell(9, "TEXT_MCHB_MATNR")
        s.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").selectedRows = "9"
        s.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").contextMenu()
        s.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").selectContextMenuItem("&XXL")

        s.findById("wnd[1]/tbar[0]/btn[0]").press()
        s.findById("wnd[1]/usr/ctxtDY_PATH").text = CAMINHO_EXPORTACAO
        s.findById("wnd[1]/usr/ctxtDY_FILENAME").text = ARQUIVO_VENC
        s.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = len(ARQUIVO_VENC)
        s.findById("wnd[1]/tbar[0]/btn[11]").press()

        time.sleep(2)
        print(f"✅ VENC exportado para {os.path.join(CAMINHO_EXPORTACAO, ARQUIVO_VENC)}")

    except Exception as e:
        print(f"❌ Erro na execução SQ00 VENC: {e}")


# ========================================
# 🧩 PÓS-PROCESSAMENTO DE PLANILHAS
# ========================================
def tratar_planilha_mb51():
    """
    Realiza o pós-processamento da planilha MB51 exportada do SAP.
    
    Remove colunas desnecessárias para otimizar o tamanho do arquivo
    e facilitar o processamento posterior no dashboard de monitoramento.
    
    Colunas removidas:
    - AA, Z, Y, X, W, U, T, S, R, Q, P, N, M, F, E, D, C, B
    
    Estas colunas contêm informações não utilizadas no dashboard ou
    dados redundantes que podem ser descartados.
    
    Raises:
        Exception: Captura e registra erros durante o processamento
        
    Note:
        O arquivo original é sobrescrito com a versão processada.
        Certifique-se de que o arquivo não está aberto em outro programa.
    """
    try:
        caminho_arquivo = os.path.join(CAMINHO_EXPORTACAO, ARQUIVO_MB51)
        print(f"🧩 Iniciando tratamento da planilha: {caminho_arquivo}")

        # Carrega o arquivo Excel
        wb = openpyxl.load_workbook(caminho_arquivo)
        ws = wb.active

        # Define colunas a serem removidas (ordem inversa para evitar problemas de índice)
        colunas_excluir = ["AA","Z","Y","X","W","U","T","S","R","Q","P","N","M","F","E","D","C","B"]

        # Remove cada coluna especificada
        for col in colunas_excluir:
            idx = openpyxl.utils.column_index_from_string(col)
            ws.delete_cols(idx)

        print(f"✅ Colunas {', '.join(colunas_excluir)} removidas com sucesso!")

        # Salva as alterações e fecha o arquivo
        wb.save(caminho_arquivo)
        wb.close()
        print("💾 Alterações salvas com sucesso!\n")

    except Exception as e:
        print(f"❌ Erro ao tratar planilha MB51: {e}")


def tratar_planilha_sq00():
    """
    Realiza o pós-processamento da planilha SQ00 exportada do SAP.
    
    Remove colunas desnecessárias para otimizar o arquivo e manter
    apenas as informações relevantes sobre validades de materiais.
    
    Colunas removidas: O, N, M, L, K, J, I, H
    
    Raises:
        Exception: Captura e registra erros durante o processamento
        
    Note:
        O arquivo original é sobrescrito com a versão processada.
    """
    try:
        caminho_arquivo = os.path.join(CAMINHO_EXPORTACAO, ARQUIVO_SQ00)
        print(f"🧩 Iniciando tratamento da planilha: {caminho_arquivo}")

        # Carrega o arquivo Excel
        wb = openpyxl.load_workbook(caminho_arquivo)
        ws = wb.active

        # Define colunas a serem removidas
        colunas_excluir = ['O', 'N', 'M', 'L', 'K', 'J', 'I', 'H']

        # Remove cada coluna especificada
        for col in colunas_excluir:
            idx = openpyxl.utils.column_index_from_string(col)
            ws.delete_cols(idx)

        print(f"✅ Colunas {', '.join(colunas_excluir)} removidas com sucesso!")

        # Salva as alterações e fecha o arquivo
        wb.save(caminho_arquivo)
        wb.close()
        print("💾 Alterações salvas com sucesso!\n")

    except Exception as e:
        print(f"❌ Erro ao tratar planilha SQ00: {e}")


def tratar_planilha_venc():
    """
    Realiza o pós-processamento da planilha de Vencimentos exportada do SAP.
    
    Remove colunas desnecessárias para otimizar o arquivo e manter
    apenas as informações essenciais sobre vencimentos de materiais.
    
    Colunas removidas: O, N, K, J, I, H
    
    Raises:
        Exception: Captura e registra erros durante o processamento
        
    Note:
        O arquivo original é sobrescrito com a versão processada.
    """
    try:
        caminho_arquivo = os.path.join(CAMINHO_EXPORTACAO, ARQUIVO_VENC)
        print(f"🧩 Iniciando tratamento da planilha: {caminho_arquivo}")

        # Carrega o arquivo Excel
        wb = openpyxl.load_workbook(caminho_arquivo)
        ws = wb.active

        # Define colunas a serem removidas
        colunas_excluir = ['O', 'N', 'K', 'J', 'I', 'H']

        # Remove cada coluna especificada
        for col in colunas_excluir:
            idx = openpyxl.utils.column_index_from_string(col)
            ws.delete_cols(idx)

        print(f"✅ Colunas {', '.join(colunas_excluir)} removidas com sucesso!")

        # Salva as alterações e fecha o arquivo
        wb.save(caminho_arquivo)
        wb.close()
        print("💾 Alterações salvas com sucesso!\n")

    except Exception as e:
        print(f"❌ Erro ao tratar planilha VENC: {e}")


# ========================================
# 🚀 FLUXO PRINCIPAL DE EXECUÇÃO
# ========================================
if __name__ == "__main__":
    """
    Fluxo principal de atualização de dados do SAP.
    
    Este script automatiza o processo completo de extração de dados do SAP:
    1. Fecha instâncias abertas do SAP para evitar conflitos
    2. Abre nova sessão SAP e realiza login
    3. Executa transações MB51, SQ00 e exporta dados
    4. Aguarda conclusão das exportações
    5. Fecha arquivos Excel abertos automaticamente
    6. Processa planilhas removendo colunas desnecessárias
    
    O processo é robusto e inclui múltiplas tentativas de fechamento
    de arquivos e verificações de disponibilidade antes do processamento.
    """
    # Etapa 1: Limpa ambiente fechando SAP aberto
    verificar_e_fechar_sap()
    
    # Etapa 2: Abre SAP e estabelece sessão
    session = abrir_sap_e_fazer_logon()
    
    if session:
        # Etapa 3: Executa transações e exporta dados
        executar_mb51(session)
        executar_sq00(session)
        executar_sq00_venc(session)

        # Etapa 4: Aguarda conclusão das exportações
        print("\n⏳ Aguardando SAP finalizar exportações...")
        time.sleep(5)  # Buffer de tempo para garantir conclusão das exportações
        
        # Etapa 5: Fecha arquivos Excel com estratégia de múltiplas tentativas
        print("\n🔒 Iniciando fechamento dos arquivos Excel...")
        
        max_tentativas = 3
        tentativa = 1
        
        while tentativa <= max_tentativas:
            print(f"\n🔄 Tentativa {tentativa} de {max_tentativas}")
            
            # Tenta fechar cada arquivo individualmente
            mb51_fechado = fechar_mb51()
            time.sleep(1)  # Pequeno delay entre fechamentos
            sq00_fechado = fechar_sq00()
            time.sleep(1)
            venc_fechado = fechar_venc()
            
            # Verifica se todos os arquivos foram fechados
            if mb51_fechado and sq00_fechado and venc_fechado:
                print("✅ Todos os arquivos fechados com sucesso!")
                break
            elif tentativa == max_tentativas:
                # Última tentativa: força fechamento de todos os processos Excel
                print("⚠️ Forçando fechamento do Excel após múltiplas tentativas...")
                forcar_fechar_excel()
                break
            else:
                print(f"⚠️ Alguns arquivos ainda abertos. Aguardando 3 segundos...")
                time.sleep(3)
                tentativa += 1
        
        # Etapa 6: Verifica disponibilidade dos arquivos para processamento
        print("\n🕐 Verificando disponibilidade dos arquivos...")
        caminho_mb51 = os.path.join(CAMINHO_EXPORTACAO, ARQUIVO_MB51)
        caminho_sq00 = os.path.join(CAMINHO_EXPORTACAO, ARQUIVO_SQ00)
        caminho_venc = os.path.join(CAMINHO_EXPORTACAO, ARQUIVO_VENC)
        
        mb51_disponivel = aguardar_arquivo_disponivel(caminho_mb51, timeout=30)
        sq00_disponivel = aguardar_arquivo_disponivel(caminho_sq00, timeout=30)
        venc_disponivel = aguardar_arquivo_disponivel(caminho_venc, timeout=30)
        
        if not (mb51_disponivel and sq00_disponivel and venc_disponivel):
            print("⚠️ Alguns arquivos ainda podem estar bloqueados, mas prosseguindo...")
        
        time.sleep(2)  # Pausa adicional de segurança antes do processamento
        
        # Etapa 7: Processa planilhas removendo colunas desnecessárias
        print("\n🧩 Iniciando tratamento das planilhas...")
        tratar_planilha_mb51()
        tratar_planilha_sq00()
        tratar_planilha_venc()

        print("\n🎯 Processo concluído com sucesso!")
    else:
        print("❌ Falha ao iniciar sessão SAP.")